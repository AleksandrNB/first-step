import openpyxl

wb = openpyxl.load_workbook("request.xlsx", read_only=True, data_only=True)
ws = wb['Main']

nmatr = []  # матрица из excel таблицы вкладка X
for row in ws.iter_rows(values_only=True):
    nmatr.append(row)

ws = wb['W']  # список W c вклвдки W
listw = []
for row in ws.iter_rows(values_only=True):
    listw.extend(row)

book = openpyxl.Workbook()  # запись в новый файл
sheet = book.active

for subarray in nmatr[:1]:  # добавить заголовок
    sheet.append(subarray)

b = 2
for i in range(1, len(listw) - 1):  # gi#размножить таблицу минус заголовок
    for subarray in nmatr[1:]:
        sheet.append(subarray)
        sheet[b][3].value = listw[i]  # добавить W
        b += 1
    print(f'обработка W : {i}')
print(f'Строк : {sheet.max_row}')
book.close()
book.save("request NSI.xlsx")  # сохранение
